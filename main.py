from helium import *
from parentvue import student, navigation
from openpyxl import Workbook, load_workbook
from os import path


def xlsx_worksheet_name(sheetname):
    # sheet name less than 35 characters
    sheetname = sheetname[:35]

    # sheet name cannot contain [ ] : * ? / \
    for ch in ['[', ']', ':', '?', '/', '\\']:
        if ch in sheetname:
            sheetname = sheetname.replace(ch, "-")

    # sheet name cannot begin or end with an apostrophe
    sheetname = sheetname.strip('\'')

    return sheetname


main_url = 'https://synergyweb.pusd11.net/'
# login_url = 'https://synergyweb.pusd11.net/PXP2_Login_Parent.aspx'
username = ''
password = ''
ws_columnname = 'Status'

# start browser and login
navigation.login(main_url, username, password)

# gather initial student data
students = student.get_students()

# gather class schedule and grades
for student in students:
    # workbook
    dest_filename = f'{student.name}.xlsx'
    if path.exists(dest_filename):
        # append to existing workbook (if exists)
        wb = load_workbook(filename=dest_filename)
    else:
        # new workbook (remove default sheet)
        wb = Workbook()
        wb.remove(wb.active)

    # gather schedule data
    student.get_classlist()

    # gather grade book data
    student.get_grades()

    # loop through classes
    for key, rows in student.grades.items():
        # worksheet for each class
        sheetname = f'{xlsx_worksheet_name(key)}'
        try:
            # append to existing worksheet (if exists)
            ws = wb[sheetname]
        except KeyError:
            # create new worksheet
            ws = wb.create_sheet(sheetname)

            # add header row
            ws.append(rows[0])
            ws.cell(row=1, column=ws.max_column + 1, value=ws_columnname)

        # append data to worksheet
        for row in rows:
            # check for existing rows
            for idx, sheetrow in enumerate(ws.iter_rows(values_only=True), start=1):
                sheetrow = tuple(str(i or '') for i in sheetrow)

                # check for duplicate rows (except last ws_columnname column)
                if row == sheetrow[0:len(sheetrow)-1]:
                    # duplicate row
                    break

                # check for matching rows
                if row[0:4] == sheetrow[0:4]:
                    # matching row found, update existing values
                    try:
                        for col, val in enumerate(row, start=1):
                            ws.cell(row=idx, column=col).value = val
                    except Exception as e:
                        print(f"Exception: {e}")

                    break
            else:
                # new row
                ws.append(row)

    # save workbook
    wb.save(filename=dest_filename)


kill_browser()
